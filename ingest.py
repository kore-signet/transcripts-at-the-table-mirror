import io
import re
import os
import json
import zipfile
from slugify import slugify
from openpyxl import load_workbook
import requests
# from requests_ratelimiter import LimiterSession
import functools
import shutil
from jinja2 import Environment, FileSystemLoader, select_autoescape
import glob

jinja_env = Environment(
    loader=FileSystemLoader("templates"), autoescape=select_autoescape()
)


formats_to_download = ["pdf", "txt", "epub"]

hyperlink_regex = re.compile(r"""^=HYPERLINK\("(.+?)",.*?"(.+?)"\)$""")
id_regex = re.compile(r"(?:id=(.+)$)|\/document\/d\/(.+?)(?:$|\/)")

# session = LimiterSession(per_second=3000)
session = requests.session()

wb = load_workbook(
    io.BytesIO(
        session.get(
            "https://docs.google.com/spreadsheets/d/1KZHwlSBvHtWStN4vTxOTrpv4Dp9WQrulwMCRocXeYcQ/export?format=xlsx",
            stream=True,
        ).content
    )
)

seasons = {}


def download_doc(episode):
    for kind in formats_to_download:
        with open(
            f"mirror/{season['id']}/{episode['slug']}.{kind}", "wb"
        ) as outf, session.get(
            f"https://docs.google.com/document/u/0/export?format={kind}&id={episode['docs_id']}",
            stream=True,
        ) as response:
            response.raw.read = functools.partial(
                response.raw.read, decode_content=True
            )
            shutil.copyfileobj(response.raw, outf)


for sheet in wb.worksheets[1:]:
    if sheet.title == "Patreon":
      continue
    season = {"title": sheet.title, "id": slugify(sheet.title), "episodes": []}
    os.makedirs(f"mirror/{season['id']}", exist_ok=True)

    ep_i = 0

    for row in sheet.iter_rows(
        min_row=(3 if 'Patreon' in sheet.title else 2), values_only=True
    ):
        if row == (None, None, None):
            continue

        hyperlink_match = hyperlink_regex.match(row[0])
        if hyperlink_match:
            title = hyperlink_match.group(2)
        else:
            title = row[0]

        title = title.strip()
        print(f"recording episode #{ep_i} - {title}")

        episode = {
            "title": title,
            "slug": slugify(title),
            "done": (row[1] or "").lower() == "yes",
            "sorting_number": ep_i,
        }

        doc_id = id_regex.search(row[2] or "")
        if doc_id:
            print(f"downloading episode #{ep_i} - {title}")
            episode["docs_id"] = doc_id.group(1) or doc_id.group(2)
            download_doc(episode)
            episode["download"] = {
                "plain": f"{season['id']}/{episode['slug']}.txt",
                "pdf": f"{season['id']}/{episode['slug']}.pdf",
                "epub": f"{season['id']}/{episode['slug']}.epub",
            }

        season["episodes"].append(episode)

        ep_i += 1

    seasons[season["id"]] = season
    with open(f"mirror/{season['id']}/index.html", "w") as season_index:
        season_index.write(
            jinja_env.get_template("season.html.jinja").render(season=season)
        )

    for ext in ['epub', 'txt']:
        with zipfile.ZipFile(f"mirror/{season['id']}-{ext}.zip", "w", compression = zipfile.ZIP_DEFLATED, compresslevel=9) as zipf:
            for path in glob.glob(f"mirror/{season['id']}/*.{ext}"):
                zipf.write(path, os.path.basename(path))

    # with open(f"mirror/{season['id']}")


with open("mirror/seasons.json", "w") as f:
    json.dump(seasons, f, indent=4)

with open("mirror/index.html", "w") as f:
    f.write(jinja_env.get_template("main.html.jinja").render(seasons=seasons))

#with open("mirror/.gitattributes", "w") as f:
#    f.write("*.zip filter=lfs diff=lfs merge=lfs -text")
    
with open("mirror/CNAME", "w") as f:
    f.write("memorious-records.cat-girl.gay")

bots_to_block = ["anthropic-ai", "CCBot", "ChatGPT-User", "FacebookBot", "GoogleOther", "Google-Extended", "GPTBot", "OmigiliBot", "Amazonbot", "Bytespider", "Claude-Web", "cohere-ai", "YouBot", "PerplexityBot", "Applebot",  "omgili", "Diffbot"]
with open("mirror/robots.txt", "w") as f:
    for bot in bots_to_block: 
        f.write(f"User-agent: {bot}\nDisallow: /\n")
